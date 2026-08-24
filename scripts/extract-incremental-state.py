#!/usr/bin/env python3
"""Extract non-secret media IDs and cutoff timestamps from an existing workbook."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SHEET_ORDER = [
    "阿飞泡枸杞",
    "欧阳会食养",
    "小七养生说",
    "肖食儿",
    "养生小禾",
    "tcmbycheehee",
    "wellness.with.gloria",
    "dr.franktcm",
    "yourtcmguide",
]
XHS_ACCOUNTS = set(SHEET_ORDER[:5])


def media_id(url: str, platform: str) -> str:
    if platform == "instagram":
        match = re.search(r"instagram\.com/(?:p|reel)/([^/?#]+)", url, re.I)
    else:
        match = re.search(r"(?:explore|profile/[^/]+)/([0-9a-f]{24})", url, re.I)
    return match.group(1) if match else ""


def iso(value) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value or "")


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract-incremental-state.py <input.xlsx> <state.json>")
    source = Path(sys.argv[1]).resolve()
    destination = Path(sys.argv[2]).resolve()
    workbook = load_workbook(source, data_only=False, read_only=True)
    if workbook.sheetnames != SHEET_ORDER:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")

    state = {"source": str(source), "accounts": {}}
    for sheet in workbook.worksheets:
        platform = "xiaohongshu" if sheet.title in XHS_ACCOUNTS else "instagram"
        url_column = 9 if platform == "xiaohongshu" else 10
        rows = []
        for row in range(2, sheet.max_row + 1):
            url = str(sheet.cell(row, url_column).value or "").strip()
            item_id = media_id(url, platform)
            if not item_id:
                continue
            rows.append({"media_id": item_id, "published_at": iso(sheet.cell(row, 1).value)})
        dated = [item for item in rows if item["published_at"]]
        latest = max(dated, key=lambda item: item["published_at"], default={})
        state["accounts"][sheet.title] = {
            "platform": platform,
            "known_media_ids": [item["media_id"] for item in rows],
            "latest_media_id": latest.get("media_id", ""),
            "latest_published_at": latest.get("published_at", ""),
            "historical_rows": len(rows),
        }

    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(destination), "accounts": {name: item["historical_rows"] for name, item in state["accounts"].items()}}, ensure_ascii=False))


if __name__ == "__main__":
    main()
