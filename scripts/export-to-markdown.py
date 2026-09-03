#!/usr/bin/env python3
"""Export legacy video-media XLSX or normalized rows.json to one Markdown file per post."""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


XHS_HEADERS = ["Timestamp", "Caption", "Script", "Likes", "Comments", "Shares", "Heat", "Duration", "URL"]
IG_HEADERS = ["Timestamp", "Caption", "Script", "Likes", "Comments", "Views", "Heat", "Engagement", "Duration", "URL"]
PLATFORM_DIRS = {
    "xiaohongshu": "xiaohongshu",
    "instagram": "instagram",
}
REQUIRED = ("Account", "Timestamp", "URL")
BAD_FILENAME = re.compile(r'[\\/:*?"<>|\x00-\x1f\x7f]+')
META_LINE = re.compile(r"^- ([A-Za-z]+):(?: (.*))?$")


def scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return str(value)
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def first_line(caption: Any) -> str:
    for line in scalar(caption).splitlines():
        if line.strip():
            return line.strip()
    return "无标题"


def safe_piece(text: str, max_bytes: int) -> str:
    value = unicodedata.normalize("NFC", text)
    value = BAD_FILENAME.sub(" ", value)
    value = re.sub(r"\s+", " ", value).strip(" .") or "无标题"
    while len(value.encode("utf-8")) > max_bytes:
        value = value[:-1].rstrip(" .")
    return value or "无标题"


def likes_piece(value: Any) -> str:
    text = scalar(value).strip()
    numeric = number(text)
    if numeric is None:
        return "点赞未知"
    return f"点赞{int(numeric) if numeric.is_integer() else text}"


def filename_for(account: str, likes: Any, title: str, suffix: int = 1) -> str:
    account_piece = safe_piece(account, 80)
    likes_label = safe_piece(likes_piece(likes), 40)
    marker = "" if suffix == 1 else f"__{suffix}"
    fixed_bytes = len(f"{account_piece} - {likes_label} - {marker}.md".encode("utf-8"))
    title_piece = safe_piece(title, max(24, 240 - fixed_bytes))
    return f"{account_piece} - {likes_label} - {title_piece}{marker}.md"


def canonical_url(value: Any) -> str:
    text = scalar(value).strip()
    if not text:
        return ""
    parts = urlsplit(text)
    host = parts.netloc.lower()
    path = parts.path.rstrip("/") + "/" if "instagram.com" in host else parts.path.rstrip("/")
    if "xiaohongshu.com" in host or "instagram.com" in host:
        return urlunsplit((parts.scheme.lower() or "https", host, path, "", ""))
    return text


def platform_for_record(record: dict[str, Any]) -> str:
    explicit = scalar(record.get("Platform")).strip().lower()
    aliases = {"xhs": "xiaohongshu", "rednote": "xiaohongshu", "ins": "instagram", "ig": "instagram"}
    explicit = aliases.get(explicit, explicit)
    if explicit in PLATFORM_DIRS:
        return explicit
    host = urlsplit(scalar(record.get("URL")).strip()).netloc.lower()
    if "xiaohongshu.com" in host:
        return "xiaohongshu"
    if "instagram.com" in host:
        return "instagram"
    raise ValueError(f"Cannot determine platform for URL: {record.get('URL', '')}")


def markdown_files(directory: Path) -> list[Path]:
    return sorted(path for path in directory.rglob("*.md") if path.is_file()) if directory.exists() else []


def headers_for_record(record: dict[str, Any]) -> list[str]:
    platform = platform_for_record(record)
    if platform == "instagram":
        return IG_HEADERS
    return XHS_HEADERS


def markdown(record: dict[str, Any], headers: list[str]) -> str:
    title = first_line(record.get("Caption"))
    metadata = ["Account", *[header for header in headers if header not in {"Caption", "Script"}]]
    lines = [f"# {title}", ""]
    lines.extend(f"- {key}: {scalar(record.get(key))}" for key in metadata)
    lines.extend(["", "## Caption", "", scalar(record.get("Caption")), "", "## Script", "", scalar(record.get("Script")), ""])
    return "\n".join(lines)


def atomic_write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            stream.write(content)
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except FileNotFoundError:
            pass
        raise


def write_records(records: Iterable[tuple[dict[str, Any], list[str]]], output_dir: Path) -> dict[str, Any]:
    existing = markdown_files(output_dir)
    used_names: defaultdict[str, set[str]] = defaultdict(set)
    seen_urls: dict[str, Path] = {}
    for path in existing:
        old = parse_record(path)
        platform = platform_for_record(old)
        used_names[platform].add(path.name)
        key = canonical_url(old.get("URL"))
        if key:
            seen_urls[key] = path
    counts: Counter[str] = Counter()
    duplicates = 0
    for record, headers in records:
        account = scalar(record.get("Account")).strip()
        url_key = canonical_url(record.get("URL"))
        if url_key and url_key in seen_urls:
            duplicates += 1
            continue
        platform = platform_for_record(record)
        suffix = 1
        while True:
            name = filename_for(account, record.get("Likes"), first_line(record.get("Caption")), suffix)
            if name not in used_names[platform]:
                break
            suffix += 1
        target = output_dir / PLATFORM_DIRS[platform] / name
        atomic_write(target, markdown(record, headers))
        used_names[platform].add(name)
        if url_key:
            seen_urls[url_key] = target
        counts[account] += 1
    return {"written": sum(counts.values()), "per_account": dict(counts), "duplicates_skipped": duplicates}


def xlsx_records(source: Path) -> Iterable[tuple[dict[str, Any], list[str]]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            first_row = next(rows, ())
            raw_headers = [scalar(value).strip() for value in first_row]
            headers = [header for header in raw_headers if header]
            if headers not in (XHS_HEADERS, IG_HEADERS):
                raise ValueError(f"Unexpected headers in {sheet.title}: {headers}")
            for values in rows:
                record = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
                if not any(record.get(key) not in (None, "") for key in ("Timestamp", "Caption", "Script", "URL")):
                    continue
                record["Account"] = sheet.title
                yield record, headers
    finally:
        workbook.close()


def number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def row_records(source: Path, existing_dir: Path | None = None) -> Iterable[tuple[dict[str, Any], list[str]]]:
    payload = json.loads(source.read_text(encoding="utf-8"))
    interactions: defaultdict[str, list[float]] = defaultdict(list)
    if existing_dir and existing_dir.exists():
        for path in markdown_files(existing_dir):
            old = parse_record(path)
            likes = number(old.get("Likes")) or 0
            comments = number(old.get("Comments")) or 0
            shares = number(old.get("Shares")) or 0
            interactions[old.get("Account", "")].append(likes + comments + shares)
    for account, rows in payload.get("accounts", {}).items():
        for row in rows:
            likes = number(row.get("likes")) or 0
            comments = number(row.get("comments")) or 0
            shares = number(row.get("shares")) or 0
            interactions[account].append(likes + comments + shares)
    for account, rows in payload.get("accounts", {}).items():
        for row in rows:
            platform = str(row.get("platform", "")).lower()
            if not platform:
                platform = platform_for_record({"URL": row.get("url")})
            is_instagram = platform == "instagram" or "views" in row
            headers = IG_HEADERS if is_instagram else XHS_HEADERS
            record = {
                "Account": account,
                "Platform": platform,
                "Timestamp": row.get("published_at"),
                "Caption": row.get("caption"),
                "Script": row.get("script"),
                "Likes": row.get("likes"),
                "Comments": row.get("comments"),
                "Duration": row.get("duration_seconds"),
                "URL": row.get("url"),
                "Heat": row.get("heat"),
            }
            likes = number(row.get("likes")) or 0
            comments = number(row.get("comments")) or 0
            shares = number(row.get("shares")) or 0
            baseline = sum(interactions[account]) / len(interactions[account]) if interactions[account] else 0
            if record["Heat"] in (None, "") and baseline:
                record["Heat"] = (likes + comments + shares) / baseline
            if is_instagram:
                views = number(row.get("views"))
                engagement = row.get("engagement")
                if engagement in (None, "") and views and views > 0:
                    engagement = (likes + comments) / views
                record.update({"Views": row.get("views"), "Engagement": engagement})
            else:
                record["Shares"] = row.get("shares")
            yield record, headers


def parse_record(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8")
    record: dict[str, str] = {}
    for line in text.splitlines():
        match = META_LINE.match(line)
        if match:
            record[match.group(1)] = match.group(2) or ""
    for field in ("Caption", "Script"):
        marker = f"## {field}\n\n"
        if marker in text:
            tail = text.split(marker, 1)[1]
            if field == "Caption":
                tail = tail.split("\n\n## Script\n\n", 1)[0]
            record[field] = tail.rstrip("\n")
    return record


def validate(output_dir: Path) -> dict[str, Any]:
    files = markdown_files(output_dir)
    errors: list[str] = []
    urls: defaultdict[str, list[str]] = defaultdict(list)
    counts: Counter[str] = Counter()
    platform_counts: Counter[str] = Counter()
    warnings: list[str] = []
    if output_dir.exists():
        for path in sorted(output_dir.rglob("*")):
            relative = path.relative_to(output_dir)
            if path.is_file() and path.suffix.lower() != ".md":
                errors.append(f"unexpected non-Markdown file: {relative}")
            if path.is_dir() and (len(relative.parts) != 1 or relative.name not in PLATFORM_DIRS.values()):
                errors.append(f"unexpected directory: {relative}")
    for path in files:
        record = parse_record(path)
        missing = [field for field in REQUIRED if not scalar(record.get(field)).strip()]
        if missing:
            errors.append(f"{path.name}: blank {', '.join(missing)}")
        optional_blanks = [field for field in ("Caption", "Script") if not scalar(record.get(field)).strip()]
        if optional_blanks:
            warnings.append(f"{path.name}: source blank {', '.join(optional_blanks)}")
        account = record.get("Account", "")
        platform = platform_for_record(record)
        platform_counts[platform] += 1
        expected_parent = PLATFORM_DIRS[platform]
        if path.parent.name != expected_parent or path.parent.parent != output_dir:
            errors.append(f"{path}: must be directly inside {expected_parent}/")
        prefix = f"{safe_piece(account, 80)} - {safe_piece(likes_piece(record.get('Likes')), 40)} - "
        if account and not path.name.startswith(prefix):
            errors.append(f"{path.name}: filename must start with account and likes")
        key = canonical_url(record.get("URL"))
        if key:
            urls[key].append(path.name)
        counts[account] += 1
    for key, names in urls.items():
        if len(names) > 1:
            errors.append(f"duplicate URL {key}: {names}")
    result = {
        "files": len(files),
        "per_platform": dict(platform_counts),
        "per_account": dict(counts),
        "errors": errors,
        "warnings": warnings,
    }
    if errors:
        raise ValueError(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def organize(source_dir: Path, output_dir: Path, remove_source: bool = False) -> dict[str, Any]:
    source_files = markdown_files(source_dir)
    if not source_files:
        raise ValueError(f"No Markdown files found beneath {source_dir}")
    records = [(parse_record(path), headers_for_record(parse_record(path))) for path in source_files]
    result = write_records(records, output_dir)
    validation = validate(output_dir)
    if validation["files"] < len(source_files):
        raise ValueError(f"Refusing cleanup: target has {validation['files']} files for {len(source_files)} sources")
    if remove_source:
        if source_dir.resolve() == output_dir.resolve():
            raise ValueError("--remove-source requires different source and output directories")
        shutil.rmtree(source_dir)
    for platform_dir in PLATFORM_DIRS.values():
        (output_dir / platform_dir).mkdir(parents=True, exist_ok=True)
    result["validation"] = validation
    result["source_removed"] = remove_source
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    for command in ("xlsx", "rows"):
        child = subparsers.add_parser(command)
        child.add_argument("source", type=Path)
        child.add_argument("output_dir", type=Path)
        child.add_argument("--existing-dir", type=Path)
    child = subparsers.add_parser("validate")
    child.add_argument("output_dir", type=Path)
    child = subparsers.add_parser("organize")
    child.add_argument("source_dir", type=Path)
    child.add_argument("output_dir", type=Path)
    child.add_argument("--remove-source", action="store_true")
    args = parser.parse_args()

    if args.command == "validate":
        result = validate(args.output_dir)
    elif args.command == "organize":
        result = organize(args.source_dir, args.output_dir, args.remove_source)
    else:
        args.output_dir.mkdir(parents=True, exist_ok=True)
        records = xlsx_records(args.source) if args.command == "xlsx" else row_records(args.source, args.existing_dir)
        result = write_records(records, args.output_dir)
        result["validation"] = validate(args.output_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
