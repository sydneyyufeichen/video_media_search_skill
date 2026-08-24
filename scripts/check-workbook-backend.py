#!/usr/bin/env python3
"""Choose a safe workbook update backend without mutating the file."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: check-workbook-backend.py <input.xlsx>")
    source = Path(sys.argv[1]).resolve()
    workbook = load_workbook(source, data_only=False, read_only=False)
    full_sheet_tables = []
    full_sheet_formulas = []
    for sheet in workbook.worksheets:
        for table in sheet.tables.values():
            if re.search(r"1048576$", str(table.ref)):
                full_sheet_tables.append({"sheet": sheet.title, "table": table.name, "ref": str(table.ref)})
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "1048576" in cell.value:
                    full_sheet_formulas.append({"sheet": sheet.title, "cell": cell.coordinate})
                    if len(full_sheet_formulas) >= 20:
                        break
            if len(full_sheet_formulas) >= 20:
                break
    requires_live = bool(full_sheet_tables or full_sheet_formulas)
    print(json.dumps({
        "source": str(source),
        "backend": "excel-live" if requires_live else "artifact-tool",
        "reason": "full-sheet table/formula ranges make offline recalculation too expensive" if requires_live else "bounded used ranges are safe for offline editing",
        "full_sheet_tables": full_sheet_tables,
        "formula_examples": full_sheet_formulas,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
