#!/usr/bin/env python3
"""Read-only structural/style audit for the canonical all-media workbook."""

from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unknown extension is not supported and will be removed")
warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported and will be removed")


def color_value(color):
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb,
        "indexed": color.indexed,
        "theme": color.theme,
        "tint": color.tint,
    }


def side_value(side):
    return {"style": side.style, "color": color_value(side.color)}


def style_value(cell):
    return {
        "style_id": cell.style_id,
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": color_value(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "fgColor": color_value(cell.fill.fgColor),
            "bgColor": color_value(cell.fill.bgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "shrink_to_fit": cell.alignment.shrink_to_fit,
        },
        "border": {
            "left": side_value(cell.border.left),
            "right": side_value(cell.border.right),
            "top": side_value(cell.border.top),
            "bottom": side_value(cell.border.bottom),
        },
    }


def conditional_formats(sheet):
    result = []
    for conditional_range, rules in sheet.conditional_formatting._cf_rules.items():
        result.append(
            {
                "range": str(conditional_range.sqref),
                "rules": [
                    {
                        "type": rule.type,
                        "priority": rule.priority,
                        "operator": rule.operator,
                        "formula": list(rule.formula or []),
                        "data_bar": {
                            "color": color_value(rule.dataBar.color),
                            "show_value": rule.dataBar.showValue,
                            "min_length": rule.dataBar.minLength,
                            "max_length": rule.dataBar.maxLength,
                        }
                        if rule.dataBar
                        else None,
                    }
                    for rule in rules
                ],
            }
        )
    return result


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: audit_workbook.py <input.xlsx> [report.json]")
    source = Path(sys.argv[1]).resolve()
    destination = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else None
    workbook = load_workbook(source, data_only=False, read_only=False)
    report = {"source": str(source), "sheet_order": workbook.sheetnames, "sheets": []}

    for sheet in workbook.worksheets:
        max_row = sheet.max_row
        max_column = sheet.max_column
        header = [sheet.cell(1, column).value for column in range(1, max_column + 1)]
        table_ranges = [str(table.ref) for table in sheet.tables.values()]
        row_heights = {
            str(row): sheet.row_dimensions[row].height
            for row in (1, 2, max_row)
            if sheet.row_dimensions[row].height is not None
        }
        styles = {}
        for row_label, row in (("header", 1), ("first_data", 2), ("last_data", max_row)):
            styles[row_label] = {
                get_column_letter(column): style_value(sheet.cell(row, column))
                for column in range(1, min(10, max_column) + 1)
            }
        report["sheets"].append(
            {
                "name": sheet.title,
                "rows": max_row,
                "data_rows": max(0, max_row - 1),
                "columns": max_column,
                "header": header,
                "tables": table_ranges,
                "auto_filter": str(sheet.auto_filter.ref or ""),
                "freeze_panes": str(sheet.freeze_panes or ""),
                "show_grid_lines": sheet.sheet_view.showGridLines,
                "column_widths": {
                    get_column_letter(column): sheet.column_dimensions[get_column_letter(column)].width
                    for column in range(1, max_column + 1)
                },
                "row_heights": row_heights,
                "conditional_formats": conditional_formats(sheet),
                "styles": styles,
                "first_date": sheet["A2"].value.isoformat() if hasattr(sheet["A2"].value, "isoformat") else sheet["A2"].value,
                "last_date": sheet.cell(max_row, 1).value.isoformat()
                if hasattr(sheet.cell(max_row, 1).value, "isoformat")
                else sheet.cell(max_row, 1).value,
                "first_metric_formulas": [sheet["I2"].value, sheet["J2"].value],
                "last_metric_formulas": [sheet.cell(max_row, 9).value, sheet.cell(max_row, 10).value],
            }
        )

    rendered = json.dumps(report, ensure_ascii=False, indent=2, default=str)
    if destination:
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(rendered, encoding="utf-8")
        print(json.dumps({
            "output": str(destination),
            "sheet_order": report["sheet_order"],
            "data_rows": {sheet["name"]: sheet["data_rows"] for sheet in report["sheets"]},
        }, ensure_ascii=False))
    else:
        print(rendered)


if __name__ == "__main__":
    main()
